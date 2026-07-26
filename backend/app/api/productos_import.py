"""Importación masiva del Maestro de Productos Karretel.

Rutas:
  POST /api/productos/import-maestro
    Lee Maestro_Productos_Karretel.xlsx desde el filesystem del servidor.
    Usada en Railway: el archivo viaja en el repositorio/imagen.

  POST /api/productos/import/karretel
    Acepta multipart/form-data con campo 'file' (.xlsx).
    Usada en desarrollo o para re-importar desde otro archivo.

Lógica compartida:
  - Crea Grupos (por Cod_Grupo), Colecciones (por nombre) y Productos (por SKU).
  - Si el SKU ya existe, actualiza precios/stock.
  - Si el SKU es nuevo, inserta el producto.
"""

import io
import os

import openpyxl
from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required

from app.extensions import db
from app.models.coleccion import Coleccion
from app.models.grupo import Grupo
from app.models.producto import Producto

productos_import_bp = Blueprint("productos_import", __name__)

# Ruta del Excel bundleado con el proyecto (Railway lo tiene en /app/)
_EXCEL_DEFAULT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "Maestro_Productos_Karretel.xlsx",
)


# ── helpers ──────────────────────────────────────────────────────────────────

def _str(val) -> str:
    return str(val).strip() if val is not None else ""


def _float_or_none(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _int_or_zero(val):
    if val is None or str(val).strip() == "":
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def _leer_excel(raw_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [
        dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
        if any(v is not None for v in row)
    ]


def _ejecutar_import(rows: list[dict]) -> dict:
    """Lógica de import compartida. Recibe filas ya parseadas, retorna reporte."""

    # Pre-carga completa para evitar autoflush conflicts durante el loop
    existing_sku: dict = {
        p.sku: p for p in Producto.query.filter(Producto.sku.isnot(None)).all()
    }
    existing_cod: dict = {p.cod_producto: p for p in Producto.query.all()}
    existing_grupos: dict = {g.cod_grupo.lower(): g for g in Grupo.query.all()}
    existing_cols: dict = {c.nombre.lower(): c for c in Coleccion.query.all()}

    errors = []
    inserted = 0
    updated = 0

    with db.session.no_autoflush:
        for idx, row in enumerate(rows, start=2):
            cod_grupo = _str(row.get("Cod_Grupo"))
            nombre_grupo = _str(row.get("Grupo"))
            nombre_coleccion = _str(row.get("Coleccion")) or nombre_grupo
            sku = _str(row.get("SKU"))
            nombre_producto = _str(row.get("Nombre_Producto"))
            precio_rollo = _float_or_none(row.get("Precio_ROLLO"))
            precio_media_rollo = _float_or_none(row.get("Precio_Media_Rollo"))
            precio_corte = _float_or_none(row.get("Precio_CORTE"))
            stock = _int_or_zero(row.get("Stock"))
            activo_raw = row.get("Activo")
            activo = bool(activo_raw) if activo_raw is not None else True

            if not sku:
                errors.append({"fila": idx, "error": "SKU vacío"})
                continue

            # ── Grupo ─────────────────────────────────────────────────────
            gk = cod_grupo.lower()
            if gk not in existing_grupos:
                g = Grupo(cod_grupo=cod_grupo, nombre=nombre_grupo)
                db.session.add(g)
                db.session.flush()
                existing_grupos[gk] = g
            grupo = existing_grupos[gk]

            # ── Coleccion ──────────────────────────────────────────────────
            ck = nombre_coleccion.lower()
            if ck not in existing_cols:
                c = Coleccion(nombre=nombre_coleccion, grupo_id=grupo.id)
                db.session.add(c)
                db.session.flush()
                existing_cols[ck] = c
            coleccion = existing_cols[ck]

            # ── Producto — deduplica por SKU ───────────────────────────────
            if sku in existing_sku:
                p = existing_sku[sku]
                p.nombre = nombre_producto or p.nombre
                p.coleccion_id = coleccion.id
                p.precio_rollo = precio_rollo
                p.precio_media_rollo = precio_media_rollo
                p.precio_corte = precio_corte
                p.stock_rollos = stock
                p.activo = activo
                updated += 1
                continue

            # Producto preexistente sin SKU cuyo cod_producto coincide con el nuevo SKU
            if sku in existing_cod:
                p = existing_cod[sku]
                if p.sku is None:
                    p.sku = sku
                    p.nombre = nombre_producto or p.nombre
                    p.coleccion_id = coleccion.id
                    p.precio_rollo = precio_rollo
                    p.precio_media_rollo = precio_media_rollo
                    p.precio_corte = precio_corte
                    p.stock_rollos = stock
                    p.activo = activo
                    existing_sku[sku] = p
                    updated += 1
                else:
                    errors.append({"fila": idx, "sku": sku,
                                   "error": "cod_producto ya pertenece a otro SKU"})
                continue

            nuevo = Producto(
                sku=sku,
                nombre=nombre_producto,
                cod_producto=sku,
                coleccion_id=coleccion.id,
                precio_rollo=precio_rollo,
                precio_media_rollo=precio_media_rollo,
                precio_corte=precio_corte,
                stock_rollos=stock,
                activo=activo,
                categoria="Telas",
                marca="Karretel",
                proveedor="Karretel",
            )
            db.session.add(nuevo)
            existing_sku[sku] = nuevo
            existing_cod[sku] = nuevo
            inserted += 1

    db.session.commit()

    return {
        "total_filas": len(rows),
        "insertados": inserted,
        "actualizados": updated,
        "errores": errors,
        "cantidad_errores": len(errors),
        "grupos_en_db": Grupo.query.count(),
        "colecciones_en_db": Coleccion.query.count(),
        "productos_en_db": Producto.query.count(),
    }


# ── rutas ─────────────────────────────────────────────────────────────────────

@productos_import_bp.post("/import-maestro")
@jwt_required()
def importar_maestro():
    """Lee el Excel desde el filesystem del servidor (Railway o local)."""
    excel_path = request.args.get("path") or _EXCEL_DEFAULT

    if not os.path.isfile(excel_path):
        return jsonify({
            "error": f"Archivo no encontrado: {excel_path}",
            "sugerencia": "Suba el archivo con POST /api/productos/import/karretel",
        }), 404

    try:
        with open(excel_path, "rb") as f:
            raw = f.read()
        rows = _leer_excel(raw)
    except Exception as exc:
        return jsonify({"error": f"No se pudo leer el archivo: {exc}"}), 400

    return jsonify(_ejecutar_import(rows))


@productos_import_bp.post("/import/karretel")
@jwt_required()
def importar_karretel():
    """Acepta el Excel como multipart/form-data campo 'file'."""
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Se requiere un archivo Excel (.xlsx)"}), 400

    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return jsonify({"error": "Formato no soportado. Use .xlsx"}), 400

    try:
        rows = _leer_excel(file.read())
    except Exception as exc:
        return jsonify({"error": f"No se pudo leer el archivo: {exc}"}), 400

    return jsonify(_ejecutar_import(rows))
