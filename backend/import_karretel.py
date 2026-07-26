"""Script de importación única del Maestro de Productos Karretel.

Uso: cd backend && python import_karretel.py
"""
import io
import sys
import openpyxl

from app import create_app
from app.extensions import db
from app.models.grupo import Grupo
from app.models.coleccion import Coleccion
from app.models.producto import Producto


def _str(val) -> str:
    return str(val).strip() if val is not None else ""


def _float_or_none(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _int_or_zero(val):
    if val is None or str(val).strip() == "":
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def run(excel_path: str = "Maestro_Productos_Karretel.xlsx"):
    app = create_app("development")
    with app.app_context():
        with open(excel_path, "rb") as f:
            raw = f.read()

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [
            dict(zip(headers, row))
            for row in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in row)
        ]
        print(f"Filas leidas: {len(rows)}")

        # Pre-load existing records to avoid autoflush conflicts
        existing_sku: dict = {p.sku: p for p in Producto.query.filter(Producto.sku.isnot(None)).all()}
        existing_cod: dict = {p.cod_producto: p for p in Producto.query.all()}
        existing_grupos: dict = {g.cod_grupo.lower(): g for g in Grupo.query.all()}
        existing_cols: dict = {c.nombre.lower(): c for c in Coleccion.query.all()}

        errors = []
        inserted = 0
        updated = 0

        with db.session.no_autoflush:
            for idx, row in enumerate(rows, start=2):
                cod_grupo = _str(row.get("Cod_Grupo"))
                nombre_grupo = _str(row.get("Grupo"))
                nombre_coleccion = _str(row.get("Coleccion")) or nombre_grupo
                sku = _str(row.get("SKU"))
                nombre_producto = _str(row.get("Nombre_Producto"))
                precio_rollo = _float_or_none(row.get("Precio_ROLLO"))
                precio_media_rollo = _float_or_none(row.get("Precio_Media_Rollo"))
                precio_corte = _float_or_none(row.get("Precio_CORTE"))
                stock = _int_or_zero(row.get("Stock"))
                activo_raw = row.get("Activo")
                activo = bool(activo_raw) if activo_raw is not None else True

                if not sku:
                    errors.append({"fila": idx, "error": "SKU vacío"})
                    continue

                # Grupo
                gk = cod_grupo.lower()
                if gk not in existing_grupos:
                    g = Grupo(cod_grupo=cod_grupo, nombre=nombre_grupo)
                    db.session.add(g)
                    db.session.flush()
                    existing_grupos[gk] = g
                grupo = existing_grupos[gk]

                # Coleccion
                ck = nombre_coleccion.lower()
                if ck not in existing_cols:
                    c = Coleccion(nombre=nombre_coleccion, grupo_id=grupo.id)
                    db.session.add(c)
                    db.session.flush()
                    existing_cols[ck] = c
                coleccion = existing_cols[ck]

                # Producto — deduplica por SKU
                if sku in existing_sku:
                    p = existing_sku[sku]
                    p.nombre = nombre_producto or p.nombre
                    p.coleccion_id = coleccion.id
                    p.precio_rollo = precio_rollo
                    p.precio_media_rollo = precio_media_rollo
                    p.precio_corte = precio_corte
                    p.stock_rollos = stock
                    p.activo = activo
                    updated += 1
                    continue

                # Manejo de colision en cod_producto (valor = sku)
                if sku in existing_cod:
                    p = existing_cod[sku]
                    if p.sku is None:
                        p.sku = sku
                        p.nombre = nombre_producto or p.nombre
                        p.coleccion_id = coleccion.id
                        p.precio_rollo = precio_rollo
                        p.precio_media_rollo = precio_media_rollo
                        p.precio_corte = precio_corte
                        p.stock_rollos = stock
                        p.activo = activo
                        existing_sku[sku] = p
                        updated += 1
                    else:
                        errors.append({"fila": idx, "sku": sku,
                                       "error": "cod_producto ya usado por otro SKU"})
                    continue

                nuevo = Producto(
                    sku=sku,
                    nombre=nombre_producto,
                    cod_producto=sku,
                    coleccion_id=coleccion.id,
                    precio_rollo=precio_rollo,
                    precio_media_rollo=precio_media_rollo,
                    precio_corte=precio_corte,
                    stock_rollos=stock,
                    activo=activo,
                    categoria="Telas",
                    marca="Karretel",
                    proveedor="Karretel",
                )
                db.session.add(nuevo)
                existing_sku[sku] = nuevo
                existing_cod[sku] = nuevo
                inserted += 1

        db.session.commit()

        print(f"Insertados:  {inserted}")
        print(f"Actualizados: {updated}")
        print(f"Errores:      {len(errors)}")
        if errors:
            for e in errors[:20]:
                print(f"  ERR: {e}")
        print(f"Grupos total:      {Grupo.query.count()}")
        print(f"Colecciones total: {Coleccion.query.count()}")
        print(f"Productos total:   {Producto.query.count()}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "Maestro_Productos_Karretel.xlsx"
    run(path)
